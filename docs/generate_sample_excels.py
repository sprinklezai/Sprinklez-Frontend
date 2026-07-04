"""
Generates sample employees.xlsx and sales_data.xlsx matching the
templates in EXCEL_TEMPLATE.md. Run once to get realistic test data,
then upload both files to your Google Drive folder.

Usage:
    pip install openpyxl --break-system-packages
    python3 generate_sample_excels.py
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook

random.seed(7)

# ---------- employees.xlsx ----------
wb = Workbook()
ws = wb.active
ws.title = "Employees"
ws.append(["username", "password", "name", "role"])
ws.append(["admin", "Admin@123", "Admin User", "Admin"])
ws.append(["jsmith", "Passw0rd!", "John Smith", "Viewer"])
ws.append(["amehta", "Sales123", "Aisha Mehta", "Viewer"])
wb.save("employees.xlsx")

# ---------- sales_data.xlsx ----------
brands = ["Brand A", "Brand B", "Brand C"]
companies = {"Brand A": "Company Alpha", "Brand B": "Company Alpha", "Brand C": "Company Beta"}
countries_regions = [
    ("UAE", "Dubai"), ("UAE", "Abu Dhabi"), ("Qatar", "Doha"),
    ("Saudi Arabia", "Riyadh"), ("Saudi Arabia", "Dammam"), ("Bahrain", "Manama"),
]
store_names = [
    "Dubai Hills", "Doha Festival City", "Doha City Centre", "Nakheel Mall - Riyadh",
    "Khaleej Mall - Riyadh", "Hamra Mall", "Bawadi Mall", "Riyadh Park Mall",
    "Nakheel Mall - Dammam", "Reem Mall", "Mall of Bahrain", "Yas Mall",
    "City Centre Ajman", "Marina Mall", "Al Ghurair Centre", "Souq Waqif",
    "Panorama Mall", "Granada Mall", "Deerah Mall", "Seef Mall",
    "Dalma Mall", "Mirdif City Centre", "Ibn Battuta Mall",
]

wb2 = Workbook()
ws_stores = wb2.active
ws_stores.title = "Stores"
ws_stores.append(["store_id", "store_name", "brand", "company", "country", "region", "status", "opened_date"])

stores = []
for i, name in enumerate(store_names, start=1):
    brand = random.choice(brands)
    country, region = random.choice(countries_regions)
    status = "Active" if random.random() > 0.1 else "Inactive"
    opened = date(2019, 1, 1) + timedelta(days=random.randint(0, 2000))
    store_id = f"ST-{i:03d}"
    stores.append((store_id, name, brand, country, region, status))
    ws_stores.append([store_id, name, brand, companies[brand], country, region, status, opened.isoformat()])

ws_brands = wb2.create_sheet("Brands")
ws_brands.append(["brand", "logo_url", "company"])
for b in brands:
    ws_brands.append([b, "", companies[b]])

ws_sales = wb2.create_sheet("Sales")
ws_sales.append(["date", "store_id", "brand", "net_revenue", "orders", "discounts", "channel"])
start = date(2026, 1, 1)
for day_offset in range(181):  # Jan - Jun 2026
    d = start + timedelta(days=day_offset)
    for store_id, name, brand, country, region, status in stores:
        if status != "Active":
            continue
        for channel, share in (("Dine-In", 0.69), ("Delivery & Takeaway", 0.31)):
            base = random.uniform(8000, 25000) * share
            net_revenue = round(base * random.uniform(0.85, 1.15), 2)
            orders = int(net_revenue / random.uniform(140, 210))
            discounts = round(net_revenue * random.uniform(0.05, 0.12), 2)
            ws_sales.append([d.isoformat(), store_id, brand, net_revenue, orders, discounts, channel])

wb2.save("sales_data.xlsx")
print("Created employees.xlsx and sales_data.xlsx in the current directory.")
